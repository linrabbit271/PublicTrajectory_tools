import os
import re
import time
import shutil
import threading
import traceback
import warnings
import pandas as pd
import openpyxl

from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QFrame, QTextEdit, QLineEdit, QProgressBar,
                             QFileDialog, QMessageBox, QWidget, QApplication, QGridLayout)
from PyQt6.QtCore import Qt, QObject, pyqtSignal
from PyQt6.QtGui import QFont, QCursor

# 自动屏蔽警告
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)


# =====================================================================
# 🌟 线程安全通讯器 (完全替代 Tkinter 的 win.after，负责前台进度与运行日志平滑刷新)
# =====================================================================
class CartonUpdater(QObject):
    log_sig = pyqtSignal(str)
    progress_sig = pyqtSignal(int, str, str)
    links_sig = pyqtSignal(str, str, str)
    finished_sig = pyqtSignal()


# =====================================================================
# 🌟 统一调用入口 (🌟 采用 *args 机制，完美净化丢弃 main.py 传过来的 self)
# =====================================================================
def open_carton_detention(*args, **kwargs):
    """主程序调用入口 (按箱号扣件)"""
    win = CartonDetentionDialog()
    win.show()


class CartonDetentionDialog(QDialog):
    def __init__(self):
        super().__init__(None)
        self._keep_alive = self  # 护盾保活，支持安全最小化与自由切屏

        self.setWindowTitle("按箱号扣件提取器")
        self.setFixedSize(850, 700)

        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint
        )
        self.setStyleSheet("background-color: #f4f6f9;")

        # 初始化路径存储
        self.keep_file_actual = ""
        self.detain_file_actual = ""

        # 挂载异步通讯槽
        self.updater = CartonUpdater()
        self.updater.log_sig.connect(self._log_slot)
        self.updater.progress_sig.connect(self._progress_slot)
        self.updater.links_sig.connect(self._links_slot)
        self.updater.finished_sig.connect(self._finished_slot)

        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ================= 顶部标题 (🌟 严格尊照核心修改 1：完美居中对齐 + 华航天空蓝企业色) =================
        header = QLabel(" 📦 报关资料 [按箱号] 扣件分流器 ")
        header.setFont(QFont("Microsoft YaHei", 12, QFont.Weight.Bold))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("background-color: #2980b9; color: white; padding: 12px 0px; border: none;")
        main_layout.addWidget(header)

        workspace = QWidget()
        work_layout = QVBoxLayout(workspace)
        work_layout.setContentsMargins(20, 15, 20, 15)
        work_layout.setSpacing(12)

        # ================= 1. 文件路径配置 =================
        lbl_cfg = QLabel("1. 文件配置")
        lbl_cfg.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        lbl_cfg.setStyleSheet("color: #34495e;")
        work_layout.addWidget(lbl_cfg)

        file_card = QFrame()
        file_card.setStyleSheet("background-color: transparent; border: none;")
        grid_lyt = QGridLayout(file_card)
        grid_lyt.setContentsMargins(0, 0, 0, 0)
        grid_lyt.setSpacing(8)

        lbl_src = QLabel("原始表单:")
        lbl_src.setFont(QFont("Microsoft YaHei", 10))
        self.entry_src = QLineEdit()
        self.entry_src.setFont(QFont("Microsoft YaHei", 10))
        self.entry_src.setStyleSheet(
            "background-color: white; border: 1px solid #bdc3c7; padding: 4px; border-radius: 2px;")
        btn_src = QPushButton("浏览...")
        btn_src.setFont(QFont("Microsoft YaHei", 9))
        btn_src.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_src.setStyleSheet(
            "background-color: #e9ecef; border: 1px solid #ced4da; padding: 4px 15px; border-radius: 2px;")
        btn_src.clicked.connect(self.select_file)

        lbl_out = QLabel("保存位置:")
        lbl_out.setFont(QFont("Microsoft YaHei", 10))
        self.entry_out = QLineEdit()
        self.entry_out.setFont(QFont("Microsoft YaHei", 10))
        self.entry_out.setStyleSheet(
            "background-color: white; border: 1px solid #bdc3c7; padding: 4px; border-radius: 2px;")
        btn_out = QPushButton("选择...")
        btn_out.setFont(QFont("Microsoft YaHei", 9))
        btn_out.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_out.setStyleSheet(
            "background-color: #e9ecef; border: 1px solid #ced4da; padding: 4px 15px; border-radius: 2px;")
        btn_out.clicked.connect(self.select_output_dir)

        grid_lyt.addWidget(lbl_src, 0, 0)
        grid_lyt.addWidget(self.entry_src, 0, 1)
        grid_lyt.addWidget(btn_src, 0, 2)
        grid_lyt.addWidget(lbl_out, 1, 0)
        grid_lyt.addWidget(self.entry_out, 1, 1)
        grid_lyt.addWidget(btn_out, 1, 2)
        grid_lyt.setColumnStretch(1, 1)
        work_layout.addWidget(file_card)

        # ================= 2. 箱号输入区 =================
        lbl_pkg = QLabel("2. 目标箱号 (支持空格、换行、逗号分隔)")
        lbl_pkg.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        lbl_pkg.setStyleSheet("color: #34495e;")
        work_layout.addWidget(lbl_pkg)

        self.text_carton_ids = QTextEdit()
        self.text_carton_ids.setFont(QFont("Consolas", 11))
        self.text_carton_ids.setStyleSheet("background-color: white; border: 1px solid #bdc3c7; border-radius: 3px;")
        self.text_carton_ids.setFixedHeight(110)
        work_layout.addWidget(self.text_carton_ids)

        # ================= 3. 操作栏与进度条 =================
        toolbar = QWidget()
        tool_layout = QHBoxLayout(toolbar)
        tool_layout.setContentsMargins(0, 0, 0, 0)
        tool_layout.setSpacing(15)

        self.btn_clear = QPushButton("🧹 清空输入")
        self.btn_run = QPushButton("🚀 开始分流提取")

        self.btn_clear.setStyleSheet(
            "QPushButton { background-color: #6c757d; color: white; border-radius: 4px; border: none; }")
        self.btn_run.setStyleSheet(
            "QPushButton { background-color: #198754; color: white; border-radius: 4px; border: none; }")

        for btn in [self.btn_clear, self.btn_run]:
            btn.setFixedHeight(40)
            btn.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

        self.btn_clear.setFixedWidth(130)
        self.btn_run.setFixedWidth(180)
        self.btn_clear.clicked.connect(self.clear_all)
        self.btn_run.clicked.connect(self.start_process_thread)

        tool_layout.addWidget(self.btn_clear)
        tool_layout.addWidget(self.btn_run)
        tool_layout.addStretch(1)
        work_layout.addWidget(toolbar)

        # 进度指示排版
        status_widget = QWidget()
        status_layout = QHBoxLayout(status_widget)
        status_layout.setContentsMargins(0, 0, 0, 0)

        self.lbl_timer = QLabel("就绪")
        self.lbl_timer.setFont(QFont("Microsoft YaHei", 9))
        self.lbl_timer.setStyleSheet("color: #7f8c8d;")

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setStyleSheet(
            "QProgressBar { background-color: #e9ecef; border: none; border-radius: 4px; } QProgressBar::chunk { background-color: #198754; border-radius: 4px; }")

        status_layout.addWidget(self.lbl_timer)
        status_layout.addWidget(self.progress_bar, stretch=1)
        work_layout.addWidget(status_widget)

        # ================= 4. 日志与快捷直达 =================
        lbl_log = QLabel("3. 运行日志与直达")
        lbl_log.setFont(QFont("Microsoft YaHei", 10, QFont.Weight.Bold))
        lbl_log.setStyleSheet("color: #34495e;")
        work_layout.addWidget(lbl_log)

        # 快捷链接容器
        self.link_frame = QWidget()
        self.link_layout = QHBoxLayout(self.link_frame)
        self.link_layout.setContentsMargins(0, 0, 0, 0)
        self.link_layout.setSpacing(12)
        self.link_layout.addStretch(1)
        work_layout.addWidget(self.link_frame)

        self.text_result_log = QTextEdit()
        self.text_result_log.setFont(QFont("Microsoft YaHei", 9))
        self.text_result_log.setStyleSheet(
            "background-color: white; border: 1px solid #bdc3c7; border-radius: 3px; color: #2c3e50;")
        self.text_result_log.setReadOnly(True)
        work_layout.addWidget(self.text_result_log, stretch=1)

        main_layout.addWidget(workspace, stretch=1)

    # ================= UI 辅助交互曹函数 =================
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择原始报关文件", "", "Excel 文件 (*.xlsx)")
        if file_path:
            self.entry_src.setText(file_path)
            if not self.entry_out.text().strip():
                self.entry_out.setText(os.path.dirname(file_path))

    def select_output_dir(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择结果保存位置")
        if dir_path:
            self.entry_out.setText(dir_path)

    def clear_all(self):
        self.entry_src.clear()
        self.text_carton_ids.clear()
        self.text_result_log.clear()
        self.progress_bar.setValue(0)
        self.lbl_timer.setText("就绪")
        self.lbl_timer.setStyleSheet("color: #7f8c8d;")

        while self.link_layout.count():
            item = self.link_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        self.link_layout.addStretch(1)

    def start_process_thread(self):
        t = threading.Thread(target=self.process_customs_data_fast, daemon=True)
        t.start()

    def open_path(self, path):
        if path and os.path.exists(path):
            os.startfile(path)

    # ================= 信号异步更新槽函数 =================
    def _log_slot(self, message):
        self.text_result_log.append(message)
        self.text_result_log.ensureCursorVisible()

    def _progress_slot(self, value, status_text, color_hex):
        self.progress_bar.setValue(value)
        self.lbl_timer.setText(status_text)
        self.lbl_timer.setStyleSheet(f"color: {color_hex};")

    def _links_slot(self, keep_path, detain_path, out_dir):
        while self.link_layout.count():
            item = self.link_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        btn_keep = QPushButton("📄 打开 [正常发运] 表")
        btn_detain = QPushButton("📦 打开 [扣件部分] 表")
        btn_dir = QPushButton("📂 打开保存目录")

        btn_keep.setStyleSheet(
            "QPushButton { background-color: #198754; color: white; font-weight: bold; padding: 5px 15px; border-radius: 2px; border: none; }")
        btn_detain.setStyleSheet(
            "QPushButton { background-color: #fd7e14; color: white; font-weight: bold; padding: 5px 15px; border-radius: 2px; border: none; }")
        btn_dir.setStyleSheet(
            "QPushButton { background-color: #6c757d; color: white; font-weight: bold; padding: 5px 15px; border-radius: 2px; border: none; }")

        for btn in [btn_keep, btn_detain, btn_dir]:
            btn.setFont(QFont("Microsoft YaHei", 9))
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

        btn_keep.clicked.connect(lambda: self.open_path(keep_path))
        btn_detain.clicked.connect(lambda: self.open_path(detain_path))
        btn_dir.clicked.connect(lambda: self.open_path(out_dir))

        self.link_layout.addWidget(btn_keep)
        self.link_layout.addWidget(btn_detain)
        self.link_layout.addWidget(btn_dir)
        self.link_layout.addStretch(1)

    def _finished_slot(self):
        self.btn_run.setEnabled(True)
        self.btn_run.setText("🚀 开始分流提取")

    # =====================================================================
    # 🌟 核心分流算法引擎 (100% 还原自动子目录归档及原样式克隆)
    # =====================================================================
    def process_customs_data_fast(self):
        src = self.entry_src.text().strip()
        out_dir = self.entry_out.text().strip()
        raw_ids = self.text_carton_ids.toPlainText().strip()

        if not src or not out_dir or not raw_ids:
            QApplication.style().staticMetaObject.invokeMethod(
                self, "warning_popup", Qt.ConnectionType.QueuedConnection
            )
            return

        # 🌟 强力净化：配合 splitlines 彻底斩断 Windows 的 \r\n 毒素
        target_ids = set([x.strip() for x in re.split(r"[\s,，\n]+", raw_ids) if x.strip()])
        if not target_ids:
            return

        start_time = time.time()
        self.btn_run.setEnabled(False)
        self.btn_run.setText("处理中...")

        self.updater.progress_sig.emit(10, "正在分析数据结构...", "#fd7e14")
        self.updater.log_sig.emit("➤ 正在匹配目标箱号...")

        try:
            # 🌟 严格尊照核心修改 2：在目标路径下全自动创建“正常发运”和“已扣件”子分类文件夹
            keep_folder = os.path.join(out_dir, "正常发运")
            detain_folder = os.path.join(out_dir, "已扣件")
            os.makedirs(keep_folder, exist_ok=True)
            os.makedirs(detain_folder, exist_ok=True)

            base_name, ext = os.path.splitext(os.path.basename(src))
            self.keep_file_actual = os.path.join(keep_folder, f"{base_name}_正常发运{ext}")
            self.detain_file_actual = os.path.join(detain_folder, f"{base_name}_已扣件{ext}")

            # 探测表头位置
            header_row_idx = 0
            for i in range(5):
                df_test = pd.read_excel(src, nrows=1, header=i)
                if any("箱号" in str(col) for col in df_test.columns):
                    header_row_idx = i
                    break

            # Pandas 高速筛选
            df_all = pd.read_excel(src, header=header_row_idx)

            try:
                carton_col = [col for col in df_all.columns if "箱号" in str(col)][0]
            except IndexError:
                self.updater.log_sig.emit("❌ 致命错误：在表格中未能找到包含“箱号”的列头！")
                self.updater.progress_sig.emit(100, "格式错误", "#dc3545")
                return

            df_all[carton_col] = df_all[carton_col].astype(str).str.strip()

            is_detain = df_all[carton_col].isin(target_ids)
            df_keep_data = df_all[~is_detain]
            df_detain_data = df_all[is_detain]
            matched_cartons = set(df_detain_data[carton_col].unique())

            self.updater.progress_sig.emit(40, "正在克隆原始表格样式...", "#fd7e14")
            self.updater.log_sig.emit("➤ 数据提取完成，正在分发至独立文件夹...")

            # 处理第一份文件 (正常发运)
            shutil.copy(src, self.keep_file_actual)
            wb_k = openpyxl.load_workbook(self.keep_file_actual)
            ws_k = wb_k.active
            data_start_row = header_row_idx + 2

            if ws_k.max_row > data_start_row + len(df_keep_data):
                ws_k.delete_rows(data_start_row + len(df_keep_data),
                                 ws_k.max_row - data_start_row - len(df_keep_data) + 1)

            for r_idx, row in enumerate(df_keep_data.values, start=data_start_row):
                for c_idx, val in enumerate(row, start=1):
                    ws_k.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            wb_k.save(self.keep_file_actual)
            wb_k.close()

            self.updater.progress_sig.emit(75, "已完成正常发运表，正在处理扣件表...", "#fd7e14")

            # 处理第二份文件 (已扣件)
            shutil.copy(src, self.detain_file_actual)
            wb_d = openpyxl.load_workbook(self.detain_file_actual)
            ws_d = wb_d.active

            if ws_d.max_row > data_start_row + len(df_detain_data):
                ws_d.delete_rows(data_start_row + len(df_detain_data),
                                 ws_d.max_row - data_start_row - len(df_detain_data) + 1)

            for r_idx, row in enumerate(df_detain_data.values, start=data_start_row):
                for c_idx, val in enumerate(row, start=1):
                    ws_d.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            wb_d.save(self.detain_file_actual)
            wb_d.close()

            elapsed = time.time() - start_time
            self.updater.progress_sig.emit(100, f"完成 (耗时 {elapsed:.2f}s)", "#198754")
            self.updater.links_sig.emit(self.keep_file_actual, self.detain_file_actual, out_dir)

            # 输出运行日志
            self.updater.log_sig.emit(f"✅ 处理完毕，数据已分别归档至独立子文件夹。")
            self.updater.log_sig.emit(f"  - 正常发运表: {len(df_keep_data)} 行")
            self.updater.log_sig.emit(f"  - 已扣留箱号: {len(df_detain_data)} 行")

            missing_ids = target_ids - matched_cartons
            if missing_ids:
                self.updater.log_sig.emit(f"\n⚠️ 警告: 有 {len(missing_ids)} 个箱号在原表中未找到:")
                self.updater.log_sig.emit(", ".join(missing_ids))

        except Exception as e:
            self.updater.log_sig.emit(f"❌ 运行错误: {str(e)}\n{traceback.format_exc()}")
            self.updater.progress_sig.emit(100, "处理异常", "#dc3545")
        finally:
            self.updater.finished_sig.emit()

    def warning_popup(self):
        QMessageBox.warning(self, "提示", "请完整填写文件路径和目标箱号！")

    def closeEvent(self, event):
        self._keep_alive = None
        event.accept()